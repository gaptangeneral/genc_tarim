from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib.messages.views import SuccessMessageMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import timedelta, date
from decimal import Decimal
from django.views.generic import TemplateView
from datetime import timedelta
from .models import CreditAccount, CreditTransaction
from .forms import CreditAccountForm, CreditTransactionForm, PaymentForm
from customers.models import Customer


class CreditAccountListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Cari hesap listesi"""
    model = CreditAccount
    template_name = 'current_accounts/account_list.html'
    context_object_name = 'accounts'
    permission_required = 'current_accounts.view_creditaccount'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('customer')
        
        # Arama filtresi
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(customer__first_name__icontains=search) |
                Q(customer__last_name__icontains=search) |
                Q(customer__company_name__icontains=search)
            )
        
        # Durum filtresi
        status = self.request.GET.get('status')
        if status == 'over_limit':
            queryset = [acc for acc in queryset if acc.is_over_limit]
        elif status == 'blocked':
            queryset = queryset.filter(is_blocked=True)
        elif status == 'active':
            queryset = queryset.filter(is_active=True, is_blocked=False)
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Cari Hesaplar"
        
        # Özet istatistikler
        total_accounts = CreditAccount.objects.count()
        total_balance = CreditAccount.objects.aggregate(
            total=Sum('current_balance')
        )['total'] or 0
        
        over_limit_count = len([
            acc for acc in CreditAccount.objects.all() 
            if acc.is_over_limit
        ])
        
        context['stats'] = {
            'total_accounts': total_accounts,
            'total_balance': total_balance,
            'over_limit_count': over_limit_count,
            'blocked_count': CreditAccount.objects.filter(is_blocked=True).count()
        }
        
        return context


class CreditAccountDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """Cari hesap detayı"""
    model = CreditAccount
    template_name = 'current_accounts/account_detail.html'
    context_object_name = 'account'
    permission_required = 'current_accounts.view_creditaccount'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        account = self.get_object()
        
        context['page_title'] = f"Cari Hesap - {account.customer}"
        
        # Son hareketler
        context['recent_transactions'] = account.transactions.all()[:20]
        
        # Ödeme formu
        context['payment_form'] = PaymentForm()
        
        # Vade analizleri
        today = timezone.now().date()
        context['overdue_transactions'] = account.transactions.filter(
            transaction_type='SALE',
            is_paid=False,
            due_date__lt=today
        )
        
        context['upcoming_payments'] = account.transactions.filter(
            transaction_type='SALE',
            is_paid=False,
            due_date__gte=today,
            due_date__lte=today + timedelta(days=30)
        )
        
        return context


class CreateCreditAccountView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    """Yeni cari hesap oluştur"""
    model = CreditAccount
    form_class = CreditAccountForm
    template_name = 'current_accounts/account_form.html'
    permission_required = 'current_accounts.add_creditaccount'
    success_message = "Cari hesap başarıyla oluşturuldu."
    success_url = reverse_lazy('current_accounts:account_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Yeni Cari Hesap"
        return context


class UpdateCreditAccountView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    """Cari hesap güncelle"""
    model = CreditAccount
    form_class = CreditAccountForm
    template_name = 'current_accounts/account_form.html'
    permission_required = 'current_accounts.change_creditaccount'
    success_message = "Cari hesap başarıyla güncellendi."
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f"Cari Hesap Düzenle - {self.object.customer}"
        return context
    
    def get_success_url(self):
        return reverse_lazy('current_accounts:account_detail', kwargs={'pk': self.object.pk})


# AJAX Views

def ajax_create_payment(request):
    """AJAX ile ödeme kaydet"""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            account_id = request.POST.get('account_id')
            amount = Decimal(request.POST.get('amount', '0'))
            description = request.POST.get('description', '')
            
            account = get_object_or_404(CreditAccount, pk=account_id)
            
            # Ödeme işlemi oluştur
            transaction = CreditTransaction.objects.create(
                credit_account=account,
                transaction_type='PAYMENT',
                amount=amount,  # Pozitif değer (alacak)
                description=description,
                created_by=request.user,
                is_paid=True,
                payment_date=timezone.now()
            )
            
            return JsonResponse({
                'success': True,
                'message': f'₺{amount} tutarında ödeme kaydedildi.',
                'new_balance': float(account.current_balance),
                'transaction_id': transaction.id
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Hata: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Geçersiz istek'})


def ajax_customer_search(request):
    """Müşteri arama (cari hesabı olmayan müşteriler)"""
    term = request.GET.get('term', '')
    
    # Cari hesabı olmayan müşterileri getir
    customers = Customer.objects.filter(
        Q(first_name__icontains=term) |
        Q(last_name__icontains=term) |
        Q(company_name__icontains=term)
    ).exclude(
        id__in=CreditAccount.objects.values_list('customer_id', flat=True)
    )[:10]
    
    results = []
    for customer in customers:
        results.append({
            'id': customer.id,
            'text': str(customer),
            'phone': customer.phone_number or '',
            'type': customer.get_customer_type_display()
        })
    
    return JsonResponse({'results': results})


def ajax_account_stats(request):
    """Dashboard için cari hesap istatistikleri"""
    today = timezone.now().date()
    
    # Vadesi geçmiş borçlar
    overdue_amount = CreditTransaction.objects.filter(
        transaction_type='SALE',
        is_paid=False,
        due_date__lt=today
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Yaklaşan vadeler (30 gün)
    upcoming_amount = CreditTransaction.objects.filter(
        transaction_type='SALE',
        is_paid=False,
        due_date__gte=today,
        due_date__lte=today + timedelta(days=30)
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Toplam bakiye
    total_balance = CreditAccount.objects.aggregate(
        total=Sum('current_balance')
    )['total'] or 0
    
    return JsonResponse({
        'overdue_amount': float(abs(overdue_amount)),  # Mutlak değer
        'upcoming_amount': float(abs(upcoming_amount)),
        'total_balance': float(total_balance),
        'account_count': CreditAccount.objects.count()
    })


# Sales uygulamasına entegrasyon için sinyal
from django.db.models.signals import post_save
from django.dispatch import receiver
from sales.models import Sale

@receiver(post_save, sender=Sale)
def create_credit_transaction_for_sale(sender, instance, created, **kwargs):
    """Satış kaydedildiğinde otomatik cari hesap hareketi oluştur"""
    
    # Sadece veresiye satışlar için
    if (instance.payment_method == 'CREDIT' and 
        instance.status == 'COMPLETED' and 
        hasattr(instance.customer, 'credit_account')):
        
        # Zaten bu satış için hareket oluşturulmuş mu kontrol et
        existing = CreditTransaction.objects.filter(
            related_sale=instance
        ).exists()
        
        if not existing:
            # Vade tarihi hesapla (30 gün sonra)
            due_date = timezone.now().date() + timedelta(days=30)
            
            CreditTransaction.objects.create(
                credit_account=instance.customer.credit_account,
                transaction_type='SALE',
                amount=-instance.grand_total,  # Negatif (borç)
                related_sale=instance,
                description=f'Satış #{instance.id} - Veresiye',
                due_date=due_date,
                created_by=instance.salesperson,
                is_paid=False
            )
            
            
def ajax_customer_credit_info(request):
    """Müşterinin cari hesap bilgilerini AJAX ile getir"""
    customer_id = request.GET.get('customer_id')
    
    if not customer_id:
        return JsonResponse({'error': 'Müşteri ID gerekli'}, status=400)
    
    try:
        customer = Customer.objects.get(pk=customer_id)
        
        if hasattr(customer, 'credit_account'):
            account = customer.credit_account
            return JsonResponse({
                'has_account': True,
                'current_balance': float(account.current_balance),
                'credit_limit': float(account.credit_limit),
                'available_credit': float(account.available_credit),
                'is_over_limit': account.is_over_limit,
                'is_blocked': account.is_blocked,
                'can_purchase': lambda amount: account.can_make_purchase(amount)  # Bu çalışmayacak, frontend'de kontrol edilmeli
            })
        else:
            return JsonResponse({
                'has_account': False,
                'current_balance': 0,
                'credit_limit': 0,
                'available_credit': 0,
                'is_over_limit': False,
                'is_blocked': False
            })
            
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Müşteri bulunamadı'}, status=404)
    
    
@login_required
def ajax_create_transaction(request):
    """AJAX ile yeni hareket kaydet"""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            account_id = request.POST.get('account_id')
            transaction_type = request.POST.get('transaction_type')
            amount = Decimal(request.POST.get('amount', '0'))
            description = request.POST.get('description', '')
            due_date = request.POST.get('due_date')
            
            account = get_object_or_404(CreditAccount, pk=account_id)
            
            # Düzeltme işlemi için negatif tutarları da kabul et
            if transaction_type == 'ADJUSTMENT':
                # Tutar olduğu gibi kaydedilir (pozitif veya negatif olabilir)
                pass
            elif transaction_type == 'PAYMENT':
                # Ödeme her zaman pozitif olmalı
                amount = abs(amount)
            
            # Vade tarihi formatı
            due_date_obj = None
            if due_date:
                try:
                    due_date_obj = timezone.datetime.strptime(due_date, '%Y-%m-%d').date()
                except ValueError:
                    pass
            
            # Hareket oluştur
            transaction = CreditTransaction.objects.create(
                credit_account=account,
                transaction_type=transaction_type,
                amount=amount,
                description=description,
                due_date=due_date_obj,
                created_by=request.user,
                is_paid=(transaction_type == 'PAYMENT'),
                payment_date=timezone.now() if transaction_type == 'PAYMENT' else None
            )
            
            return JsonResponse({
                'success': True,
                'message': f'{transaction.get_transaction_type_display()} başarıyla kaydedildi.',
                'new_balance': float(account.current_balance),
                'transaction_id': transaction.id
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Hata: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Geçersiz istek'})


def ajax_customer_credit_info(request):
    """Müşterinin cari hesap bilgilerini AJAX ile getir"""
    customer_id = request.GET.get('customer_id')
    
    if not customer_id:
        return JsonResponse({'error': 'Müşteri ID gerekli'}, status=400)
    
    try:
        customer = Customer.objects.get(pk=customer_id)
        
        if hasattr(customer, 'credit_account'):
            account = customer.credit_account
            return JsonResponse({
                'has_account': True,
                'current_balance': float(account.current_balance),
                'credit_limit': float(account.credit_limit),
                'available_credit': float(account.available_credit),
                'is_over_limit': account.is_over_limit,
                'is_blocked': account.is_blocked,
            })
        else:
            return JsonResponse({
                'has_account': False,
                'current_balance': 0,
                'credit_limit': 0,
                'available_credit': 0,
                'is_over_limit': False,
                'is_blocked': False
            })
            
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Müşteri bulunamadı'}, status=404)
    
    
class CreditAccountReportsView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Cari hesap raporları sayfası"""
    template_name = 'current_accounts/account_reports.html'
    permission_required = 'current_accounts.view_creditaccount'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Cari Hesap Raporları"
        
        # Rapor filtreleri
        report_type = self.request.GET.get('report_type', 'summary')
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        customer_type = self.request.GET.get('customer_type')
        
        # Tarih aralığı belirleme
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        end_date = timezone.now()
        start_date = end_date - timedelta(days=30)  # Varsayılan 30 gün
        
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                start_date = timezone.make_aware(start_date) if timezone.is_naive(start_date) else start_date
            except ValueError:
                pass
        
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
                # Günün sonuna kadar dahil etmek için
                end_date = end_date.replace(hour=23, minute=59, second=59)
                end_date = timezone.make_aware(end_date) if timezone.is_naive(end_date) else end_date
            except ValueError:
                pass
        
        # Temel veriler
        accounts_queryset = CreditAccount.objects.select_related('customer').all()
        
        # Müşteri tipi filtresi
        if customer_type:
            accounts_queryset = accounts_queryset.filter(customer__customer_type=customer_type)
        
        # Risk analizi ile birlikte hesaplamalar
        accounts_with_analysis = []
        for account in accounts_queryset:
            risk_percent = 0
            risk_status = 'low'
            
            if account.credit_limit > 0:
                # Risk yüzdesi hesaplama: Kullanılan limitin yüzdesi
                used_limit = account.credit_limit + account.current_balance  # current_balance negatifse kullanılan limit artar
                risk_percent = max(0, (used_limit / account.credit_limit) * 100)
                
                if risk_percent >= 90:
                    risk_status = 'high'
                elif risk_percent >= 70:
                    risk_status = 'medium'
            
            accounts_with_analysis.append({
                'account': account,
                'risk_percent': risk_percent,
                'risk_status': risk_status
            })
        
        context['accounts'] = accounts_with_analysis
        
        # Özet istatistikler - Tarih filtreli
        today = timezone.now().date()
        
        # Toplam alacaklar (negatif bakiyeler)
        total_receivables = accounts_queryset.filter(current_balance__lt=0).aggregate(
            total=Sum('current_balance')
        )['total'] or 0
        total_receivables = abs(total_receivables)
        
        # Vadesi geçen borçlar - tarih filtreli
        overdue_transactions = CreditTransaction.objects.filter(
            transaction_type='SALE',
            is_paid=False,
            due_date__lt=today,
            created_at__range=(start_date, end_date)
        )
        
        overdue_amount = overdue_transactions.aggregate(
            total=Sum('amount')
        )['total'] or 0
        overdue_amount = abs(overdue_amount)
        
        # 30 gün içinde vadesi gelecek - tarih filtreli
        upcoming_transactions = CreditTransaction.objects.filter(
            transaction_type='SALE',
            is_paid=False,
            due_date__gte=today,
            due_date__lte=today + timedelta(days=30),
            created_at__range=(start_date, end_date)
        )
        upcoming_amount = upcoming_transactions.aggregate(
            total=Sum('amount')
        )['total'] or 0
        upcoming_amount = abs(upcoming_amount)
        
        # Yaşlandırma analizi - tarih filtreli
        aging_data = {
            'current': 0,  # Vadesi gelmemiş
            'days_1_30': 0,  # 1-30 gün gecikmiş
            'days_31_60': 0,  # 31-60 gün gecikmiş
            'days_over_60': 0,  # 60+ gün gecikmiş
        }
        
        # Vadesi geçmiş işlemleri yaşlarına göre kategorize et
        for transaction in overdue_transactions:
            if transaction.due_date:
                days_overdue = (today - transaction.due_date).days
                amount = abs(transaction.amount)
                
                if days_overdue <= 30:
                    aging_data['days_1_30'] += amount
                elif days_overdue <= 60:
                    aging_data['days_31_60'] += amount
                else:
                    aging_data['days_over_60'] += amount
        
        # Vadesi gelmemiş işlemler - tarih filtreli
        current_transactions = CreditTransaction.objects.filter(
            transaction_type='SALE',
            is_paid=False,
            due_date__gte=today,
            created_at__range=(start_date, end_date)
        )
        aging_data['current'] = abs(current_transactions.aggregate(
            total=Sum('amount')
        )['total'] or 0)
        
        context.update({
            'total_receivables': total_receivables,
            'overdue_amount': overdue_amount,
            'upcoming_amount': upcoming_amount,
            'aging': aging_data,
            'report_type': report_type,
            'start_date': start_date,
            'end_date': end_date,
            'start_date_str': start_date_str or '',
            'end_date_str': end_date_str or '',
            'customer_type': customer_type or '',
        })
        
        return context
    
@login_required
def export_accounts_excel(request):
    """Cari hesapları Excel'e aktar"""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill
    
    # Workbook oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cari Hesaplar"
    
    # Başlık satırı
    headers = [
        'Müşteri Adı', 
        'Müşteri Tipi', 
        'Güncel Bakiye (₺)', 
        'Kredi Limiti (₺)', 
        'Kullanılabilir Kredi (₺)',
        'Durum',
        'Oluşturma Tarihi',
        'Son Güncelleme'
    ]
    
    # Başlık stilini ayarla
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Veri satırları
    accounts = CreditAccount.objects.select_related('customer').all()
    
    for row, account in enumerate(accounts, 2):
        ws.cell(row=row, column=1, value=str(account.customer))
        ws.cell(row=row, column=2, value=account.customer.get_customer_type_display())
        ws.cell(row=row, column=3, value=float(account.current_balance))
        ws.cell(row=row, column=4, value=float(account.credit_limit))
        ws.cell(row=row, column=5, value=float(account.available_credit))
        
        # Durum
        if account.is_blocked:
            status = "Bloke"
        elif account.is_over_limit:
            status = "Limit Aştı"
        elif account.is_active:
            status = "Aktif"
        else:
            status = "Pasif"
        
        ws.cell(row=row, column=6, value=status)
        ws.cell(row=row, column=7, value=account.created_at.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=8, value=account.updated_at.strftime('%d.%m.%Y %H:%M'))
    
    # Sütun genişliklerini ayarla
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # HTTP response olarak döndür
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=cari_hesaplar.xlsx'
    
    wb.save(response)
    return response